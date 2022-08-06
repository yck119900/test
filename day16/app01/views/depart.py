from django.shortcuts import render, redirect,HttpResponse
from app01 import models
from app01.utils.pagination import Pagination


def depart_list(request):
    """部门列表"""
    # 去数据库中获取所有部门列表
    queryset = models.Department.objects.all()
    page_object = Pagination(request, queryset, page_size=2)
    content = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html()
    }

    return render(request, 'depart_list.html', content)


def depart_add(request):
    """添加部门"""
    # 如果是GET请求，那么就直接显示添加部门页面
    if request.method == 'GET':
        return render(request, 'depart_add.html')
    # 获取用户POST提交过来的数据
    title = request.POST.get('title')
    models.Department.objects.create(title=title)
    # 重定向回部门列表
    return redirect('/depart/list/')


def depart_delete(request):
    """删除部门"""
    # 获取id
    nid = request.GET.get('nid')
    # 删除
    models.Department.objects.filter(id=nid).delete()
    # 重定向回部门列表
    return redirect('/depart/list/')


def depart_edit(request, nid):
    """编辑部门"""
    if request.method == 'GET':
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html', {'row_object': row_object})
    title = request.POST.get('title')
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect('/depart/list/')

def depart_multi(request):
    """批量上传excel"""
    from openpyxl import load_workbook
    file_object=request.FILES.get('exc')
    wb=load_workbook(file_object)
    sheet=wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        text=row[0].value
        exists=models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')

